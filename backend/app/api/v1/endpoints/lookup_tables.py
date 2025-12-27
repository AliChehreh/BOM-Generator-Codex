from io import BytesIO
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.api.deps import get_db
from app.models.lookup_table import LookupTable
from app.models.enums import LookupTableScope
from app.models.lookup_table_column import LookupTableColumn
from app.models.lookup_table_row import LookupTableRow
from app.schemas.lookup_table import (
    LookupTableColumnsUpdate,
    LookupTableCreate,
    LookupTableOut,
    LookupTableColumnOut,
    LookupTableRowBulkUpdate,
    LookupTableRowOut,
    LookupTableUpdate,
)

router = APIRouter()


@router.get("/lookup-tables", response_model=list[LookupTableOut])
def list_lookup_tables(
    scope: LookupTableScope | None = None,
    build_family_id: UUID | None = None,
    db: Session = Depends(get_db),
):
    stmt = select(LookupTable)
    if scope:
        stmt = stmt.where(LookupTable.scope == scope)
    if build_family_id:
        stmt = stmt.where(LookupTable.build_family_id == build_family_id)
    return db.execute(stmt.order_by(LookupTable.name)).scalars().all()


@router.post("/lookup-tables", response_model=LookupTableOut, status_code=status.HTTP_201_CREATED)
def create_lookup_table(payload: LookupTableCreate, db: Session = Depends(get_db)):
    if payload.scope == LookupTableScope.build_family and payload.build_family_id is None:
        raise HTTPException(status_code=400, detail="build_family_id required for build_family scope")
    table = LookupTable(
        name=payload.name,
        scope=payload.scope,
        build_family_id=payload.build_family_id,
        description=payload.description,
    )
    db.add(table)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Lookup table name already exists in scope")
    db.refresh(table)
    return table


@router.get("/lookup-tables/{table_id}", response_model=LookupTableOut)
def get_lookup_table(table_id: UUID, db: Session = Depends(get_db)):
    table = db.get(LookupTable, table_id)
    if not table:
        raise HTTPException(status_code=404, detail="Lookup table not found")
    return table


@router.put("/lookup-tables/{table_id}", response_model=LookupTableOut)
def update_lookup_table(
    table_id: UUID, payload: LookupTableUpdate, db: Session = Depends(get_db)
):
    table = db.get(LookupTable, table_id)
    if not table:
        raise HTTPException(status_code=404, detail="Lookup table not found")
    if payload.scope == LookupTableScope.build_family and payload.build_family_id is None:
        raise HTTPException(status_code=400, detail="build_family_id required for build_family scope")
    table.name = payload.name
    table.scope = payload.scope
    table.build_family_id = payload.build_family_id
    table.description = payload.description
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Lookup table name already exists in scope")
    db.refresh(table)
    return table


@router.delete("/lookup-tables/{table_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_lookup_table(table_id: UUID, db: Session = Depends(get_db)):
    table = db.get(LookupTable, table_id)
    if not table:
        raise HTTPException(status_code=404, detail="Lookup table not found")
    db.delete(table)
    db.commit()
    return None


@router.put("/lookup-tables/{table_id}/columns")
def update_lookup_columns(
    table_id: UUID, payload: LookupTableColumnsUpdate, db: Session = Depends(get_db)
):
    table = db.get(LookupTable, table_id)
    if not table:
        raise HTTPException(status_code=404, detail="Lookup table not found")

    existing = (
        db.execute(
            select(LookupTableColumn).where(LookupTableColumn.lookup_table_id == table_id)
        )
        .scalars()
        .all()
    )
    existing_by_id = {col.id: col for col in existing}
    payload_ids = {col.id for col in payload.columns if col.id}

    for col in existing:
        if col.id not in payload_ids:
            db.delete(col)

    for col in payload.columns:
        if col.id and col.id in existing_by_id:
            existing_col = existing_by_id[col.id]
            existing_col.column_name = col.column_name
            existing_col.column_type = col.column_type
            existing_col.display_order = col.display_order
        else:
            db.add(
                LookupTableColumn(
                    lookup_table_id=table_id,
                    column_name=col.column_name,
                    column_type=col.column_type,
                    display_order=col.display_order,
                )
            )

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Duplicate column name in lookup table")

    return {"status": "ok"}


@router.get("/lookup-tables/{table_id}/columns", response_model=list[LookupTableColumnOut])
def list_lookup_columns(table_id: UUID, db: Session = Depends(get_db)):
    return (
        db.execute(
            select(LookupTableColumn)
            .where(LookupTableColumn.lookup_table_id == table_id)
            .order_by(LookupTableColumn.display_order, LookupTableColumn.column_name)
        )
        .scalars()
        .all()
    )


@router.get("/lookup-tables/{table_id}/rows", response_model=list[LookupTableRowOut])
def list_lookup_rows(table_id: UUID, db: Session = Depends(get_db)):
    return (
        db.execute(
            select(LookupTableRow)
            .where(LookupTableRow.lookup_table_id == table_id)
            .order_by(LookupTableRow.list_size)
        )
        .scalars()
        .all()
    )


@router.put("/lookup-tables/{table_id}/rows", response_model=list[LookupTableRowOut])
def upsert_lookup_rows(
    table_id: UUID, payload: LookupTableRowBulkUpdate, db: Session = Depends(get_db)
):
    table = db.get(LookupTable, table_id)
    if not table:
        raise HTTPException(status_code=404, detail="Lookup table not found")

    existing = (
        db.execute(
            select(LookupTableRow).where(LookupTableRow.lookup_table_id == table_id)
        )
        .scalars()
        .all()
    )
    existing_by_size = {float(row.list_size): row for row in existing}

    for row in payload.rows:
        key = float(row.list_size)
        if key in existing_by_size:
            existing_row = existing_by_size[key]
            existing_row.row_values_json = row.row_values_json
        else:
            db.add(
                LookupTableRow(
                    lookup_table_id=table_id,
                    list_size=row.list_size,
                    row_values_json=row.row_values_json,
                )
            )

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Duplicate list_size in lookup table")

    return (
        db.execute(
            select(LookupTableRow)
            .where(LookupTableRow.lookup_table_id == table_id)
            .order_by(LookupTableRow.list_size)
        )
        .scalars()
        .all()
    )


@router.post("/lookup-tables/{table_id}/import")
def import_lookup_rows(
    table_id: UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    table = db.get(LookupTable, table_id)
    if not table:
        raise HTTPException(status_code=404, detail="Lookup table not found")

    columns = (
        db.execute(
            select(LookupTableColumn)
            .where(LookupTableColumn.lookup_table_id == table_id)
            .order_by(LookupTableColumn.display_order)
        )
        .scalars()
        .all()
    )
    def normalize_header(value: str) -> str:
        return value.strip().lower().replace(" ", "_").replace("-", "_")

    column_lookup = {normalize_header(col.column_name): col.column_name for col in columns}

    contents = file.file.read()
    rows = []
    filename = file.filename or ""
    if filename.lower().endswith(".csv"):
        import csv

        reader = csv.DictReader(contents.decode("utf-8-sig").splitlines())
        for row in reader:
            normalized_row = {}
            for key, value in row.items():
                if key is None:
                    continue
                normalized_row[normalize_header(str(key))] = value
            rows.append(normalized_row)
    elif filename.lower().endswith(".xlsx"):
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(contents), data_only=True)
        sheet = workbook.active
        headers = []
        for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
            raw_header = "" if cell.value is None else str(cell.value)
            headers.append(normalize_header(raw_header))
        for row in sheet.iter_rows(min_row=2, values_only=True):
            normalized_row = {}
            for idx, value in enumerate(row):
                if idx >= len(headers):
                    continue
                header = headers[idx]
                if not header:
                    continue
                normalized_row[header] = value
            rows.append(normalized_row)
    else:
        raise HTTPException(status_code=400, detail="Only CSV or XLSX files are supported")

    parsed_rows = []
    for row_index, row in enumerate(rows, start=2):
        if not any(value not in (None, "") for value in row.values()):
            continue
        if "list_size" not in row:
            raise HTTPException(
                status_code=400,
                detail="Missing list_size column. Expected a 'list_size' header.",
            )
        raw_size = row.get("list_size")
        if raw_size in (None, ""):
            raise HTTPException(
                status_code=400,
                detail=f"Missing list_size value on row {row_index}",
            )
        try:
            list_size = float(raw_size)
        except (TypeError, ValueError):
            raise HTTPException(
                status_code=400,
                detail=f"Invalid list_size value on row {row_index}",
            )
        values = {}
        for normalized_name, original_name in column_lookup.items():
            if normalized_name in row and normalized_name != "list_size":
                values[original_name] = row[normalized_name]
        parsed_rows.append({"list_size": list_size, "row_values_json": values})

    payload = LookupTableRowBulkUpdate(rows=parsed_rows)
    return upsert_lookup_rows(table_id, payload, db)
